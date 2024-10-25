# filename: gen_wallets_with_images.py

import os
import qrcode
from eth_account import Account
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter

# Define o diretório onde os QR codes serão salvos
wallet_dir = "wallets"
os.makedirs(wallet_dir, exist_ok=True)

# Cria um arquivo Excel (.xlsx)
wb = Workbook()
ws = wb.active
ws.title = "Ethereum Wallets"

# Adiciona cabeçalhos
ws.append(["Wallet Address/PubKey", "Public Key QR Code", "Private Key", "Private Key QR Code"])

# Define a largura das colunas e a altura das linhas para acomodar os QR codes
for col in range(2, 5):  # As colunas que armazenam QR codes
    col_letter = get_column_letter(col)
    ws.column_dimensions[col_letter].width = 30  # Ajuste da largura da coluna

# Define a altura das linhas
row_height = 120  # Altura suficiente para o QR code
ws.row_dimensions[1].height = row_height  # Para a linha do cabeçalho

# Gera 100 carteiras Ethereum (pode ser ajustado para 1000)
for i in range(1, 10):
    # Gera a conta Ethereum
    account = Account.create()

    # Extrai chaves pública e privada
    pub_address = account.address
    priv_key = account.key.hex()

    # Gera o QR code para o endereço público
    pub_qr_img = qrcode.make(pub_address)
    pub_qr_path = os.path.join(wallet_dir, f"{pub_address}_PUB.png")
    pub_qr_img.save(pub_qr_path)

    # Gera o QR code para a chave privada
    priv_qr_img = qrcode.make(priv_key)
    priv_qr_path = os.path.join(wallet_dir, f"{pub_address}_PRIV.png")
    priv_qr_img.save(priv_qr_path)

    # Adiciona os dados da carteira ao Excel
    ws.append([pub_address, '', priv_key, ''])

    # Ajusta a altura da linha corrente para acomodar o QR code
    ws.row_dimensions[i + 1].height = row_height

    # Insere o QR code da chave pública na célula correspondente
    pub_img = ExcelImage(pub_qr_path)
    pub_img.width = 100  # Largura do QR code
    pub_img.height = 100  # Altura do QR code
    ws.add_image(pub_img, f"B{i + 1}")  # Coloca na coluna B (chave pública)

    # Insere o QR code da chave privada na célula correspondente
    priv_img = ExcelImage(priv_qr_path)
    priv_img.width = 100
    priv_img.height = 100
    ws.add_image(priv_img, f"D{i + 1}")  # Coloca na coluna D (chave privada)

    # Exibe progresso
    print(f"Carteira {i} gerada: {pub_address}")

# Salva o arquivo Excel
excel_file = os.path.join(wallet_dir, "ethereum_wallets_with_qrcodes.xlsx")
wb.save(excel_file)

print(f"Arquivo Excel salvo com QR codes: {excel_file}")